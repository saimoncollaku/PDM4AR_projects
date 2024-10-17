from abc import ABC, abstractmethod
from turtle import st

import numpy as np
from numpy.typing import NDArray
from traitlets import Bool, Int
from pdm4ar.exercises.ex04.structures import Action, Policy, State, ValueFunc, Cell
from typing import Dict, Tuple

import openpyxl
from itertools import product
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


class GridMdp:
    def __init__(self, grid: NDArray[np.int64], gamma: float = 0.9):
        assert len(grid.shape) == 2, "Map is invalid"
        self.grid = grid
        self.gamma: float = gamma
        self.prob_wormhole = 1 / np.sum(self.grid == Cell.WORMHOLE)
        # self.max = max(grid.flatten())
        self.rows, self.cols = grid.shape

        self.transition_data: Dict[Tuple[State, Action], Dict[State, float]] = {}
        self.reward_data: Dict[Tuple[State, Action], Dict[State, float]] = {}
        self.admissible_actions_data: Dict[State, list[Action]] = {}
        self.admissible_next_states_data: Dict[State, list[State]] = {}
        self.start_state = list(zip(*np.where(self.grid == Cell.START)))
        self.wormhole_states = list(zip(*np.where(self.grid == Cell.WORMHOLE)))

    def get_admissible_actions(self, state: State) -> list[Action]:
        if state not in self.admissible_actions_data:
            cell = self.grid[state]
            if cell == Cell.CLIFF:
                actions = [Action.ABANDON]
            elif cell == Cell.GOAL:
                actions = [Action.STAY]
            else:
                actions = self.get_admitted_directions(state)
                actions.append(Action.ABANDON)
            self.admissible_actions_data[state] = actions
        return self.admissible_actions_data[state]

    def get_admissible_next_states(self, state: State) -> list[State]:
        if state not in self.admissible_next_states_data:
            cell = self.grid[state]
            next_states = []
            if cell == Cell.CLIFF:
                next_states = self.start_state
            elif cell == Cell.GOAL:
                next_states = [state]
            else:
                admitted_dir = self.get_admitted_directions(state)
                if self.check_adj_to_wormhole(state, admitted_dir[0], admitted_dir):
                    next_states += self.wormhole_states
                elif self.check_dir_to_wormhole(state, admitted_dir[0], admitted_dir):
                    next_states += self.wormhole_states
                if Action.NORTH in admitted_dir:
                    next_states.append((state[0] - 1, state[1]))
                if Action.SOUTH in admitted_dir:
                    next_states.append((state[0] + 1, state[1]))
                if Action.EAST in admitted_dir:
                    next_states.append((state[0], state[1] + 1))
                if Action.WEST in admitted_dir:
                    next_states.append((state[0], state[1] - 1))
                if admitted_dir != 4:
                    next_states += self.start_state
                if cell == Cell.SWAMP:
                    next_states.append(state)
                next_states = list(set(next_states))
            self.admissible_next_states_data[state] = next_states
        return self.admissible_next_states_data[state]

    def get_transition_prob(self, state: State, action: Action, next_state: State) -> float:
        state_action_key = (state, action)
        if state_action_key not in self.transition_data:
            self.transition_data[state_action_key] = {}
            current_cell = self.grid[state]
            if current_cell == Cell.CLIFF:
                self.transition_data[state_action_key][next_state] = 0
            else:
                prob_fun = {
                    Cell.GRASS: self.compute_prob_grass,
                    Cell.WORMHOLE: self.compute_prob_grass,
                    Cell.START: self.compute_prob_grass,
                    Cell.SWAMP: self.compute_prob_swamp,
                    Cell.GOAL: self.compute_prob_goal,
                }.get(current_cell, lambda *args: 0)
                for possible_next_state in self.get_admissible_next_states(state):
                    prob = prob_fun(state, action, possible_next_state)
                    self.transition_data[state_action_key][possible_next_state] = prob
        return self.transition_data[state_action_key].get(next_state, 0)

    def stage_reward(self, state: State, action: Action, next_state: State) -> float:
        state_action_key = (state, action)
        if state_action_key not in self.reward_data:
            self.reward_data[state_action_key] = {}
            current_cell = self.grid[state]
            if current_cell == Cell.CLIFF:
                self.reward_data[state_action_key][next_state] = 0
            else:
                reward_func = {
                    Cell.GRASS: self.compute_reward_grass,
                    Cell.WORMHOLE: self.compute_reward_grass,
                    Cell.START: self.compute_reward_grass,
                    Cell.SWAMP: self.compute_reward_swamp,
                    Cell.GOAL: self.compute_reward_goal,
                }.get(current_cell, lambda *args: 0)
                for possible_next_state in self.get_admissible_next_states(state):
                    reward = reward_func(state, action, possible_next_state)
                    self.reward_data[state_action_key][possible_next_state] = reward
        return self.reward_data[state_action_key].get(next_state, 0)

    def compute_reward_grass(self, state: State, action: Action, next_state: State) -> float:
        next_cell = self.grid[next_state]
        admitted_dir = self.get_admitted_directions(state)

        if next_cell == Cell.CLIFF:
            return 0
        if action == Action.ABANDON:
            return -10 if next_cell == Cell.START else 0
        if next_cell == Cell.START:
            return -11 if len(admitted_dir) != 4 else -1
        if next_cell == Cell.WORMHOLE:
            if self.check_adj_to_wormhole(state, action, admitted_dir) != 0:
                return -1
            if self.check_dir_to_wormhole(state, action, admitted_dir):
                return -1
            return 0
        if self.check_next_state_is_adj(state, next_state, admitted_dir):
            return -1
        return 0

    def compute_reward_swamp(self, state: State, action: Action, next_state: State) -> float:
        next_cell = self.grid[next_state]
        admitted_dir = self.get_admitted_directions(state)

        if next_cell == Cell.CLIFF:
            return 0
        if action == Action.ABANDON:
            return -10 if next_cell == Cell.START else 0
        if self.check_same_state(state, next_state) and action != Action.ABANDON:
            return -2
        if next_cell == Cell.START:
            return -12 if len(admitted_dir) != 4 else -2
        if next_cell == Cell.WORMHOLE:
            if self.check_adj_to_wormhole(state, action, admitted_dir) != 0:
                return -2
            if self.check_dir_to_wormhole(state, action, admitted_dir):
                return -2
            return 0
        if self.check_next_state_is_adj(state, next_state, admitted_dir):
            return -2
        return 0

    def compute_reward_goal(self, state: State, action: Action, next_state: State) -> float:

        if action == Action.STAY and state == next_state:
            return 50
        return 0

    def compute_prob_grass(self, state: State, action: Action, next_state: State) -> float:
        next_cell = self.grid[next_state]
        admitted_dir = self.get_admitted_directions(state)

        if action == Action.ABANDON:
            return 1 if next_cell == Cell.START else 0

        if next_cell == Cell.START:
            prob_cliff = (4 - len(admitted_dir)) * (0.25 / 3)
            if self.check_next_state_is_dir(state, action, next_state, admitted_dir):
                return 0.75 + prob_cliff
            if self.check_next_state_is_adj(state, next_state, admitted_dir):
                return (0.25 / 3) + prob_cliff
            if action not in admitted_dir:
                return 0.75 + prob_cliff - (0.25 / 3)
            return prob_cliff

        if next_cell == Cell.WORMHOLE:
            p = 0
            if self.check_dir_to_wormhole(state, action, admitted_dir):
                p += 0.75 * self.prob_wormhole
            p += self.check_adj_to_wormhole(state, action, admitted_dir) * ((0.25 / 3)) * self.prob_wormhole
            return p

        if self.check_next_state_is_dir(state, action, next_state, admitted_dir):
            return 0.75
        if self.check_next_state_is_adj(state, next_state, admitted_dir):
            return 0.25 / 3

        return 0

    def compute_prob_swamp(self, state: State, action: Action, next_state: State) -> float:
        next_cell = self.grid[next_state]
        admitted_dir = self.get_admitted_directions(state)

        if action == Action.ABANDON:
            return 1 if next_cell == Cell.START else 0
        if action == Action.STAY:
            return 1 if self.check_same_state(state, next_state) else 0

        prob_cliff = (4 - len(admitted_dir)) * (0.25 / 3) + 0.05

        if next_cell == Cell.START:
            if self.check_next_state_is_dir(state, action, next_state, admitted_dir):
                return 0.5 + prob_cliff
            if self.check_next_state_is_adj(state, next_state, admitted_dir):
                return (0.25 / 3) + prob_cliff
            if action not in admitted_dir:
                return 0.5 + prob_cliff - (0.25 / 3)
            return prob_cliff

        if next_cell == Cell.WORMHOLE:
            p = 0
            if self.check_dir_to_wormhole(state, action, admitted_dir):
                p += 0.5 * self.prob_wormhole
            p += self.check_adj_to_wormhole(state, action, admitted_dir) * ((0.25 / 3)) * self.prob_wormhole
            return p

        if self.check_next_state_is_dir(state, action, next_state, admitted_dir):
            return 0.5
        if self.check_next_state_is_adj(state, next_state, admitted_dir):
            return 0.25 / 3
        if self.check_same_state(state, next_state):
            return 0.2

        return 0

    def compute_prob_goal(self, state: State, action: Action, next_state: State) -> float:

        if action == Action.STAY and state == next_state:
            return 1
        return 0

    def get_admitted_directions(self, state: State) -> list[Action]:
        directions = []
        for action, (dr, dc) in zip(
            [Action.SOUTH, Action.NORTH, Action.EAST, Action.WEST], [(1, 0), (-1, 0), (0, 1), (0, -1)]
        ):
            new_r, new_c = state[0] + dr, state[1] + dc
            if 0 <= new_r < self.rows and 0 <= new_c < self.cols and self.grid[new_r, new_c] != Cell.CLIFF:
                directions.append(action)
        return directions

    def check_same_state(self, state: State, next_state: State) -> bool:
        return state == next_state

    def check_dir_to_wormhole(self, state: State, action: Action, admissible_dir: list[Action]) -> bool:
        if action not in admissible_dir:
            return False
        dr, dc = {Action.NORTH: (-1, 0), Action.SOUTH: (1, 0), Action.EAST: (0, 1), Action.WEST: (0, -1)}[action]
        return self.grid[state[0] + dr, state[1] + dc] == Cell.WORMHOLE

    def check_adj_to_wormhole(self, state: State, action: Action, admitted_dir: list[Action]) -> int:
        return sum(
            1
            for a in admitted_dir
            if a != action
            and self.grid[
                state[0] + {Action.NORTH: -1, Action.SOUTH: 1, Action.EAST: 0, Action.WEST: 0}[a],
                state[1] + {Action.NORTH: 0, Action.SOUTH: 0, Action.EAST: 1, Action.WEST: -1}[a],
            ]
            == Cell.WORMHOLE
        )

    def check_next_state_is_dir(
        self, state: State, action: Action, next_state: State, admitted_dir: list[Action]
    ) -> bool:
        if action not in admitted_dir:
            return False
        dr, dc = {Action.NORTH: (-1, 0), Action.SOUTH: (1, 0), Action.EAST: (0, 1), Action.WEST: (0, -1)}[action]
        return next_state == (state[0] + dr, state[1] + dc)

    def check_next_state_is_adj(self, state: State, next_state: State, admitted_dir: list[Action]) -> bool:
        return any(self.check_next_state_is_dir(state, action, next_state, admitted_dir) for action in admitted_dir)

    def create_excel_file(self, filename: str = "grid_mdp_data.xlsx"):
        # Create a new workbook and select the active sheet
        wb = openpyxl.Workbook()

        # Create sheets for probabilities and rewards
        prob_sheet = wb.active
        prob_sheet.title = "Transition Probabilities"
        reward_sheet = wb.create_sheet(title="Rewards")

        # Get all possible states and actions
        rows, cols = self.grid.shape
        states = list(product(range(rows), range(cols)))
        actions = list(Action)

        # Create header row with all possible next states
        header = ["State_Action"] + [f"{s[0]},{s[1]}" for s in states]

        # Write headers to both sheets
        for sheet in [prob_sheet, reward_sheet]:
            sheet.append(header)
            for cell in sheet[1]:
                cell.font = Font(bold=True)

        # Populate data for each state-action combination
        for state in states:
            for action in actions:
                state_action = f"{state[0]},{state[1]}_{action.name}"
                prob_row = [state_action]
                reward_row = [state_action]

                for next_state in states:
                    prob = self.get_transition_prob(state, action, next_state)
                    reward = self.stage_reward(state, action, next_state)

                    prob_row.append(prob)
                    reward_row.append(reward)

                prob_sheet.append(prob_row)
                reward_sheet.append(reward_row)

        # Adjust column widths
        for sheet in [prob_sheet, reward_sheet]:
            for column in sheet.columns:
                max_length = 0
                column_letter = openpyxl.utils.get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = max_length + 2
                sheet.column_dimensions[column_letter].width = adjusted_width

        # Save the workbook
        wb.save(filename)

        print(f"Excel file has been created: {filename}")

    def compute_all_action(self):
        for state in np.ndindex(self.grid.shape):
            self.get_admissible_actions(state)
            self.get_admissible_next_states(state)


class GridMdpSolver(ABC):
    @staticmethod
    @abstractmethod
    def solve(grid_mdp: GridMdp) -> tuple[ValueFunc, Policy]:
        pass
